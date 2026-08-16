from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

wb = Workbook()
ink = "1B1814"
muted = "6F675C"
paper = "F3EEE4"
card = "FFFCF6"
accent = "B4451B"
done = "2C5A4A"
line = "D8D0C2"
blue = "0000FF"

thin = Border(
    left=Side(style="thin", color=line),
    right=Side(style="thin", color=line),
    top=Side(style="thin", color=line),
    bottom=Side(style="thin", color=line),
)
head_font = Font(name="Arial", bold=True, color="FFF8EE", size=11)
title_font = Font(name="Arial", bold=True, color=ink, size=16)
label_font = Font(name="Arial", color=ink, size=11)
muted_font = Font(name="Arial", color=muted, size=10)
input_font = Font(name="Arial", color=blue, size=11)
formula_font = Font(name="Arial", color="000000", size=11)
head_fill = PatternFill("solid", fgColor=ink)
paper_fill = PatternFill("solid", fgColor=paper)
card_fill = PatternFill("solid", fgColor=card)
yellow = PatternFill("solid", fgColor="FFF3BF")
ok_fill = PatternFill("solid", fgColor="E4EEE8")
warn_fill = PatternFill("solid", fgColor="F6DDD3")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = thin

# --- 가설 ---
ws = wb.active
ws.title = "가설"
ws.sheet_properties.tabColor = ink
ws["A1"] = "오늘마감 KPI 검증표"
ws["A1"].font = title_font
ws.merge_cells("A1:D1")
ws["A2"] = "CHS7006 실습 4 · 이재훈 2021314509 · https://oneulmagam.vercel.app"
ws["A2"].font = muted_font
ws.merge_cells("A2:D2")

rows = [
    ("항목", "내용"),
    ("가설", "오늘·이번 주 마감만 한 화면에서 보면, 매일 아이캠퍼스 전체를 훑지 않아도 마감을 놓치지 않는다."),
    ("페르소나", "한 학기 5~7과목, 아이캠퍼스를 매일 확인하는 습관이 있고, 마감 놓침을 학점 손실로 느끼는 학부생"),
    ("성공 기준 1", "1주 마감 놓침 0건"),
    ("성공 기준 2", "아이캠퍼스 접속 주 7회 → 주 2회 이하"),
    ("성공 기준 3", "아침 1회 확인 후 오늘 항목을 모두 처리함으로 표시 (today_cleared)"),
    ("설문 대신 볼 것", "visit / complete_item / today_cleared / share_click / import_feed"),
]
ws["A4"] = rows[0][0]
ws["B4"] = rows[0][1]
style_header(ws, 4, 2)
for i, (a, b) in enumerate(rows[1:], start=5):
    ws[f"A{i}"] = a
    ws[f"B{i}"] = b
    ws[f"A{i}"].font = Font(name="Arial", bold=True, size=11)
    ws[f"B{i}"].font = label_font
    ws[f"A{i}"].fill = paper_fill
    ws[f"B{i}"].fill = card_fill
    ws[f"A{i}"].alignment = left
    ws[f"B{i}"].alignment = left
    ws[f"A{i}"].border = thin
    ws[f"B{i}"].border = thin

ws["A13"] = "노란 칸은 매일 저녁에 직접 적는다. 파란 글씨는 입력, 검은 글씨는 수식이다."
ws["A13"].font = muted_font
ws.merge_cells("A13:D13")
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 78
ws.row_dimensions[1].height = 24
for r in range(5, 12):
    ws.row_dimensions[r].height = 28

# --- 일지 ---
log = wb.create_sheet("일지")
log.sheet_properties.tabColor = accent
log["A1"] = "7일 검증 일지"
log["A1"].font = title_font
log.merge_cells("A1:H1")
headers = ["날짜", "요일", "아이캠퍼스 접속 횟수", "오늘마감 사용", "today_cleared", "마감 놓침", "메모", "판정"]
for i, h in enumerate(headers, 1):
    log.cell(3, i, h)
style_header(log, 3, 8)

weekdays = ["월", "화", "수", "목", "금", "토", "일"]
for i, day in enumerate(weekdays):
    r = 4 + i
    log.cell(r, 1, f"2026-08-{17+i:02d}").font = input_font
    log.cell(r, 2, day).font = formula_font
    log.cell(r, 3, 1 if i < 2 else 0).font = input_font
    log.cell(r, 4, "Y").font = input_font
    log.cell(r, 5, "Y" if i != 2 else "N").font = input_font
    log.cell(r, 6, 0).font = input_font
    log.cell(r, 7, "").font = input_font
    log.cell(r, 8, f'=IF(AND(C{r}<=1,F{r}=0),"통과","미달")')
    log.cell(r, 8).font = formula_font
    for c in range(1, 9):
        log.cell(r, c).border = thin
        log.cell(r, c).alignment = center
        if c in (3, 4, 5, 6, 7):
            log.cell(r, c).fill = yellow
        if c == 8:
            log.cell(r, c).fill = card_fill

log.conditional_formatting.add(
    "H4:H10",
    FormulaRule(formula=['$H4="통과"'], fill=ok_fill),
)
log.conditional_formatting.add(
    "H4:H10",
    FormulaRule(formula=['$H4="미달"'], fill=warn_fill),
)

for i, w in enumerate([14, 8, 20, 16, 16, 12, 28, 10], 1):
    log.column_dimensions[get_column_letter(i)].width = w
log["A12"] = "노란 칸만 바꾼다. 판정 열은 수식이다. 오늘마감 사용 / today_cleared 는 Y 또는 N."
log["A12"].font = muted_font
log.merge_cells("A12:H12")

# --- 집계 ---
summ = wb.create_sheet("집계")
summ.sheet_properties.tabColor = done
summ["A1"] = "주간 집계"
summ["A1"].font = title_font
summ.merge_cells("A1:C1")
summ["A3"] = "지표"
summ["B3"] = "값"
summ["C3"] = "기준"
style_header(summ, 3, 3)

metrics = [
    ("아이캠퍼스 접속 합계", "=SUM(일지!C4:C10)", "주 2회 이하"),
    ("오늘마감 사용 일수", '=COUNTIF(일지!D4:D10,"Y")', "7일 중 5일 이상"),
    ("today_cleared 일수", '=COUNTIF(일지!E4:E10,"Y")', "사용한 날의 대부분"),
    ("마감 놓침 합계", "=SUM(일지!F4:F10)", "0건"),
    ("통과 일수", '=COUNTIF(일지!H4:H10,"통과")', "5일 이상"),
    ("가설 채택", '=IF(AND(B4<=2,B7=0,B8>=5),"채택","기각 또는 재실험")', "접속↓ 그리고 놓침 0"),
]
for i, (name, formula, rule) in enumerate(metrics, 4):
    summ.cell(i, 1, name).font = Font(name="Arial", bold=True, size=11)
    summ.cell(i, 2, formula).font = formula_font
    summ.cell(i, 3, rule).font = muted_font
    for c in range(1, 4):
        summ.cell(i, c).border = thin
        summ.cell(i, c).alignment = left if c != 2 else center
        summ.cell(i, c).fill = card_fill
    summ.cell(i, 2).fill = paper_fill

summ["A11"] = "이벤트 로그(사이트에서 KPI 내려받기)를 붙일 때"
summ["A11"].font = Font(name="Arial", bold=True, size=11)
summ["A12"] = "visit 횟수"
summ["B12"] = '=COUNTIF(이벤트로그!A:A,"visit")'
summ["A13"] = "today_cleared 횟수"
summ["B13"] = '=COUNTIF(이벤트로그!A:A,"today_cleared")'
summ["A14"] = "share_click 횟수"
summ["B14"] = '=COUNTIF(이벤트로그!A:A,"share_click")'
summ["A15"] = "import_feed 횟수"
summ["B15"] = '=COUNTIF(이벤트로그!A:A,"import_feed")'
for r in range(12, 16):
    summ.cell(r, 1).font = label_font
    summ.cell(r, 2).font = formula_font
    summ.cell(r, 1).border = thin
    summ.cell(r, 2).border = thin

summ.column_dimensions["A"].width = 28
summ.column_dimensions["B"].width = 36
summ.column_dimensions["C"].width = 28

chart = BarChart()
chart.type = "col"
chart.title = "일별 아이캠퍼스 접속"
chart.y_axis.title = "횟수"
chart.x_axis.title = None
data = Reference(log, min_col=3, min_row=3, max_row=10)
cats = Reference(log, min_col=2, min_row=4, max_row=10)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.style = 10
chart.y_axis.scaling.min = 0
chart.legend = None
chart.width = 15
chart.height = 8
summ.add_chart(chart, "A18")

# --- 이벤트로그 ---
ev = wb.create_sheet("이벤트로그")
ev.sheet_properties.tabColor = "3D4F7C"
ev["A1"] = "event_name"
ev["B1"] = "item_type"
ev["C1"] = "used_at"
style_header(ev, 1, 3)
ev["A2"] = "visit"
ev["B2"] = ""
ev["C2"] = "사이트의 KPI 내려받기 CSV를 여기부터 붙여넣는다"
ev["A2"].font = muted_font
ev["C2"].font = muted_font
ev.column_dimensions["A"].width = 20
ev.column_dimensions["B"].width = 16
ev.column_dimensions["C"].width = 48
ev["A2"].fill = yellow
ev["B2"].fill = yellow
ev["C2"].fill = yellow

out = "/Users/leejaehun/Desktop/오늘마감/오늘마감_KPI.xlsx"
wb.save(out)
print(out)
