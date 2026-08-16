from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


wb = Workbook()
INK, MUTED, LINE = "1B1814", "6F675C", "D8D0C2"
PAPER, CARD, ACCENT, DONE = "F3EEE4", "FFFCF6", "B4451B", "2C5A4A"
BLUE, YELLOW = "0000FF", "FFF3BF"
thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
head_f = Font(name="Arial", bold=True, color="FFF8EE", size=11)
title_f = Font(name="Arial", bold=True, color=INK, size=16)
ink_f = Font(name="Arial", color=INK, size=11)
muted_f = Font(name="Arial", color=MUTED, size=10)
in_f = Font(name="Arial", color=BLUE, size=11)
form_f = Font(name="Arial", color="000000", size=11)
head_fill = PatternFill("solid", fgColor=INK)
paper_fill = PatternFill("solid", fgColor=PAPER)
card_fill = PatternFill("solid", fgColor=CARD)
yellow = PatternFill("solid", fgColor=YELLOW)
ok_fill = PatternFill("solid", fgColor="E4EEE8")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def paint_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = head_f
        cell.fill = head_fill
        cell.alignment = center
        cell.border = thin


def box(cell, font=ink_f, fill=card_fill, align=left):
    cell.font = font
    cell.fill = fill
    cell.alignment = align
    cell.border = thin


# ---------- 담당자마스터 ----------
emp = wb.active
emp.title = "담당자마스터"
emp.sheet_properties.tabColor = DONE
emp["A1"] = "담당자 마스터"
emp["A1"].font = title_f
emp.merge_cells("A1:F1")
emp["A2"] = "실습 4 · CHS7006 · 입력(파란 글씨)은 마스터, 실습 시트는 수식만 쓴다"
emp["A2"].font = muted_f
headers = ["담당자ID", "이름", "소속팀", "직급", "지역", "월목표"]
for i, h in enumerate(headers, 1):
    emp.cell(4, i, h)
paint_header(emp, 4, 6)
staff = [
    ("EMP001", "김서연", "서울1팀", "사원", "서울", 18000000),
    ("EMP002", "박민준", "서울1팀", "대리", "서울", 24000000),
    ("EMP003", "이하늘", "부산팀", "과장", "부산", 30000000),
    ("EMP004", "최유진", "대구팀", "사원", "대구", 16000000),
    ("EMP005", "정우성", "광주팀", "대리", "광주", 22000000),
    ("EMP006", "한지민", "인천팀", "과장", "인천", 28000000),
    ("EMP013", "황성민", "대전팀", "팀장", "대전", 35500000),
    ("EMP018", "오세훈", "서울2팀", "사원", "서울", 15000000),
]
for i, row in enumerate(staff):
    r = 5 + i
    for c, val in enumerate(row, 1):
        cell = emp.cell(r, c, val)
        box(cell, in_f, yellow, center if c != 2 else left)
        if c == 6:
            cell.number_format = '#,##0'
for i, w in enumerate([12, 12, 12, 10, 10, 14], 1):
    emp.column_dimensions[get_column_letter(i)].width = w

# ---------- 상품마스터 ----------
prd = wb.create_sheet("상품마스터")
prd.sheet_properties.tabColor = ACCENT
prd["A1"] = "상품 마스터"
prd["A1"].font = title_f
ph = ["상품ID", "상품명", "카테고리", "기준단가", "원가", "마진율", "재고"]
for i, h in enumerate(ph, 1):
    prd.cell(4, i, h)
paint_header(prd, 4, 7)
products = [
    ("P101", "캠퍼스 노트", "문구", 4500, 2200, 80),
    ("P102", "형광펜 세트", "문구", 8900, 4100, 120),
    ("P201", "무선 마우스", "디지털", 24900, 13200, 40),
    ("P202", "기계식 키보드", "디지털", 89000, 51000, 18),
    ("P203", "USB-C 허브", "디지털", 32900, 17800, 35),
    ("P301", "텀블러 500ml", "생활", 18900, 8200, 55),
    ("P302", "에코백", "생활", 12900, 4800, 90),
    ("P401", "스터디 램프", "가구", 45900, 24100, 22),
    ("P402", "접이식 책상", "가구", 129000, 72000, 12),
    ("P501", "비타민C", "헬스", 15900, 6100, 70),
]
prd["A2"] = "마진율은 수식 = (기준단가-원가)/기준단가"
prd["A2"].font = muted_f
for i, (pid, name, cat, price, cost, stock) in enumerate(products):
    r = 5 + i
    for c, val in enumerate((pid, name, cat, price, cost), 1):
        box(prd.cell(r, c, val), in_f, yellow, center if c != 2 else left)
        if c in (4, 5):
            prd.cell(r, c).number_format = '#,##0'
    box(prd.cell(r, 6, f"=(D{r}-E{r})/D{r}"), form_f, card_fill, center)
    prd.cell(r, 6).number_format = "0.0%"
    box(prd.cell(r, 7, stock), in_f, yellow, center)
for i, w in enumerate([12, 18, 12, 12, 12, 12, 10], 1):
    prd.column_dimensions[get_column_letter(i)].width = w

# ---------- 판매원본 ----------
sale = wb.create_sheet("판매원본", 0)
sale.sheet_properties.tabColor = INK
sale["A1"] = "판매 원본"
sale["A1"].font = title_f
sale["A2"] = "파란 글씨=입력, 검은 글씨=수식. 연/월/분기·매출·할인금액·순매출은 계산한다."
sale["A2"].font = muted_f
sh = ["주문번호", "주문일자", "연도", "월", "분기", "판매지역", "담당자ID", "상품ID",
      "판매채널", "결제수단", "수량", "단가", "할인율", "매출금액", "할인금액", "순매출금액", "상태"]
for i, h in enumerate(sh, 1):
    sale.cell(4, i, h)
paint_header(sale, 4, 17)

channels = ["온라인몰", "온라인몰", "온라인몰", "오프라인", "오프라인", "앱"]
pays = ["카드", "간편결제", "계좌이체", "카드"]
status = ["완료", "완료", "완료", "취소", "완료"]
emp_ids = [s[0] for s in staff]
prd_ids = [p[0] for p in products]
prd_price = {p[0]: p[3] for p in products}

start = date(2024, 1, 8)
n_orders = 48
for i in range(n_orders):
    r = 5 + i
    d = start + timedelta(days=(i * 7) % 350 + (i % 5))
    pid = prd_ids[i % len(prd_ids)]
    qty = 1 + (i * 3) % 8
    disc = [0, 0, 0.05, 0.1, 0][i % 5]
    # inputs
    values = {
        1: f"ORD{2024000 + i + 1}",
        2: d,
        6: staff[i % len(staff)][4],
        7: emp_ids[i % len(emp_ids)],
        8: pid,
        9: channels[i % len(channels)],
        10: pays[i % len(pays)],
        11: qty,
        12: prd_price[pid],
        13: disc,
        17: status[i % len(status)],
    }
    for c, val in values.items():
        cell = sale.cell(r, c, val)
        box(cell, in_f, yellow, center)
        if c == 2:
            cell.number_format = "YYYY-MM-DD"
        if c in (11, 12):
            cell.number_format = "#,##0"
        if c == 13:
            cell.number_format = "0%"
    # formulas
    formulas = {
        3: f"=YEAR(B{r})",
        4: f"=MONTH(B{r})",
        5: f'=IF(D{r}<=3,1,IF(D{r}<=6,2,IF(D{r}<=9,3,4)))',
        14: f"=K{r}*L{r}",
        15: f"=N{r}*M{r}",
        16: f"=N{r}-O{r}",
    }
    for c, fml in formulas.items():
        cell = sale.cell(r, c, fml)
        box(cell, form_f, card_fill, center)
        if c in (14, 15, 16):
            cell.number_format = "#,##0"
    sale.cell(r, 6).alignment = center

for i, w in enumerate([12, 13, 8, 6, 6, 10, 12, 10, 12, 12, 8, 10, 10, 12, 12, 12, 10], 1):
    sale.column_dimensions[get_column_letter(i)].width = w
last = 4 + n_orders

# ---------- 실습 ----------
ex = wb.create_sheet("실습")
ex.sheet_properties.tabColor = "3D4F7C"
ex["A1"] = "실습 워크시트"
ex["A1"].font = title_f
ex["A2"] = "강의와 같이 빈칸은 수식으로 채운다. VLOOKUP / SUMIFS / COUNTIFS / IFS."
ex["A2"].font = muted_f

# Section 1
ex["A4"] = "Section 1. 담당자ID로 이름·소속·직급·월목표"
ex["A4"].font = Font(name="Arial", bold=True, size=12, color=INK)
s1h = ["담당자ID", "이름", "소속팀", "직급", "월목표"]
for i, h in enumerate(s1h, 1):
    ex.cell(5, i, h)
paint_header(ex, 5, 5)
for i, sid in enumerate(["EMP013", "EMP001", "EMP003", "EMP018", "EMP006"]):
    r = 6 + i
    box(ex.cell(r, 1, sid), in_f, yellow, center)
    box(ex.cell(r, 2, f'=VLOOKUP(A{r},담당자마스터!$A$5:$F$12,2,FALSE)'), form_f)
    box(ex.cell(r, 3, f'=VLOOKUP(A{r},담당자마스터!$A$5:$F$12,3,FALSE)'), form_f)
    box(ex.cell(r, 4, f'=VLOOKUP(A{r},담당자마스터!$A$5:$F$12,4,FALSE)'), form_f)
    box(ex.cell(r, 5, f'=VLOOKUP(A{r},담당자마스터!$A$5:$F$12,6,FALSE)'), form_f, align=center)
    ex.cell(r, 5).number_format = "#,##0"

# Section 2
ex["A13"] = "Section 2. 상품ID로 상품명·카테고리·기준단가·마진율"
ex["A13"].font = Font(name="Arial", bold=True, size=12, color=INK)
s2h = ["상품ID", "상품명", "카테고리", "기준단가", "마진율"]
for i, h in enumerate(s2h, 1):
    ex.cell(14, i, h)
paint_header(ex, 14, 5)
for i, sid in enumerate(["P202", "P101", "P401", "P501", "P203"]):
    r = 15 + i
    box(ex.cell(r, 1, sid), in_f, yellow, center)
    box(ex.cell(r, 2, f'=VLOOKUP(A{r},상품마스터!$A$5:$G$14,2,FALSE)'), form_f)
    box(ex.cell(r, 3, f'=VLOOKUP(A{r},상품마스터!$A$5:$G$14,3,FALSE)'), form_f)
    box(ex.cell(r, 4, f'=VLOOKUP(A{r},상품마스터!$A$5:$G$14,4,FALSE)'), form_f, align=center)
    box(ex.cell(r, 5, f'=VLOOKUP(A{r},상품마스터!$A$5:$G$14,6,FALSE)'), form_f, align=center)
    ex.cell(r, 4).number_format = "#,##0"
    ex.cell(r, 5).number_format = "0.0%"

# Section 3
ex["A22"] = "Section 3. 채널=온라인몰 주문 건수·매출·순매출 (취소 제외)"
ex["A22"].font = Font(name="Arial", bold=True, size=12, color=INK)
s3h = ["조건", "주문건수", "매출합계", "순매출합계"]
for i, h in enumerate(s3h, 1):
    ex.cell(23, i, h)
paint_header(ex, 23, 4)
box(ex.cell(24, 1, "온라인몰"), in_f, yellow, center)
box(ex.cell(24, 2, f'=COUNTIFS(판매원본!$I$5:$I${last},A24,판매원본!$Q$5:$Q${last},"완료")'), form_f, align=center)
box(ex.cell(24, 3, f'=SUMIFS(판매원본!$N$5:$N${last},판매원본!$I$5:$I${last},A24,판매원본!$Q$5:$Q${last},"완료")'), form_f, align=center)
box(ex.cell(24, 4, f'=SUMIFS(판매원본!$P$5:$P${last},판매원본!$I$5:$I${last},A24,판매원본!$Q$5:$Q${last},"완료")'), form_f, align=center)
ex.cell(24, 3).number_format = "#,##0"
ex.cell(24, 4).number_format = "#,##0"

# Section 4
ex["A27"] = "Section 4. 담당자 순매출 등급  500만↑ VIP / 200~499만 GOLD / 100~199만 SILVER / 그 외 BRONZE"
ex["A27"].font = Font(name="Arial", bold=True, size=12, color=INK)
s4h = ["담당자ID", "이름", "순매출", "등급"]
for i, h in enumerate(s4h, 1):
    ex.cell(28, i, h)
paint_header(ex, 28, 4)
for i, sid in enumerate(["EMP013", "EMP002", "EMP004", "EMP001", "EMP018"]):
    r = 29 + i
    box(ex.cell(r, 1, sid), in_f, yellow, center)
    box(ex.cell(r, 2, f'=VLOOKUP(A{r},담당자마스터!$A$5:$F$12,2,FALSE)'), form_f)
    box(ex.cell(r, 3, f'=SUMIFS(판매원본!$P$5:$P${last},판매원본!$G$5:$G${last},A{r},판매원본!$Q$5:$Q${last},"완료")'), form_f, align=center)
    box(ex.cell(r, 4, f'=IFS(C{r}>=5000000,"VIP",C{r}>=2000000,"GOLD",C{r}>=1000000,"SILVER",TRUE,"BRONZE")'), form_f, align=center)
    ex.cell(r, 3).number_format = "#,##0"

ex.conditional_formatting.add("D29:D33", FormulaRule(formula=['$D29="VIP"'], fill=ok_fill))
for i, w in enumerate([14, 16, 16, 16, 14], 1):
    ex.column_dimensions[get_column_letter(i)].width = w

# ---------- 대시보드 ----------
dash = wb.create_sheet("대시보드")
dash.sheet_properties.tabColor = "9A6B12"
dash["A1"] = "월별 판매 대시보드"
dash["A1"].font = title_f
dash["A2"] = "완료 주문만. 수치는 판매원본을 참조한다."
dash["A2"].font = muted_f
for i, h in enumerate(["월", "주문건수", "매출", "순매출", "목표대비"], 1):
    dash.cell(4, i, h)
paint_header(dash, 4, 5)
for m in range(1, 13):
    r = 4 + m
    box(dash.cell(r, 1, m), form_f, paper_fill, center)
    box(dash.cell(r, 2, f'=COUNTIFS(판매원본!$D$5:$D${last},A{r},판매원본!$Q$5:$Q${last},"완료")'), form_f, align=center)
    box(dash.cell(r, 3, f'=SUMIFS(판매원본!$N$5:$N${last},판매원본!$D$5:$D${last},A{r},판매원본!$Q$5:$Q${last},"완료")'), form_f, align=center)
    box(dash.cell(r, 4, f'=SUMIFS(판매원본!$P$5:$P${last},판매원본!$D$5:$D${last},A{r},판매원본!$Q$5:$Q${last},"완료")'), form_f, align=center)
    box(dash.cell(r, 5, f'=IF(C{r}=0,"-",D{r}/SUM(담당자마스터!$F$5:$F$12))'), form_f, align=center)
    dash.cell(r, 3).number_format = "#,##0"
    dash.cell(r, 4).number_format = "#,##0"
    dash.cell(r, 5).number_format = "0.0%"

dash["A18"] = "전체 완료 순매출"
dash["B18"] = f'=SUMIF(판매원본!$Q$5:$Q${last},"완료",판매원본!$P$5:$P${last})'
dash["A19"] = "전체 완료 주문"
dash["B19"] = f'=COUNTIF(판매원본!$Q$5:$Q${last},"완료")'
dash["A20"] = "온라인몰 비중"
dash["B20"] = f'=IF(B19=0,"-",실습!B24/B19)'
dash["B18"].number_format = "#,##0"
dash["B20"].number_format = "0.0%"
for r in range(18, 21):
    dash.cell(r, 1).font = Font(name="Arial", bold=True, size=11)
    box(dash.cell(r, 2), form_f, paper_fill, center)

chart = BarChart()
chart.type = "col"
chart.title = "월별 순매출"
chart.y_axis.title = "원"
data = Reference(dash, min_col=4, min_row=4, max_row=16)
cats = Reference(dash, min_col=1, min_row=5, max_row=16)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.style = 10
chart.legend = None
chart.width = 18
chart.height = 9
dash.add_chart(chart, "G4")
for i, w in enumerate([10, 12, 14, 14, 12], 1):
    dash.column_dimensions[get_column_letter(i)].width = w

# ---------- 보고서 ----------
rep = wb.create_sheet("보고서")
rep.sheet_properties.tabColor = "5B3A6E"
rep["A1"] = "2024 판매 보고서 초안"
rep["A1"].font = title_f
rep["A3"] = "이 시트 숫자는 대시보드·실습 시트를 참조한다. 글을 고칠 때도 수식은 건드리지 않는다."
rep["A3"].font = muted_f
rep.merge_cells("A3:F3")
rep["A5"] = "핵심 숫자"
rep["A5"].font = Font(name="Arial", bold=True, size=13)
rep["A6"] = "완료 주문 수"
rep["B6"] = "=대시보드!B19"
rep["A7"] = "완료 순매출"
rep["B7"] = "=대시보드!B18"
rep["A8"] = "온라인몰 주문"
rep["B8"] = "=실습!B24"
rep["A9"] = "온라인몰 순매출"
rep["B9"] = "=실습!D24"
rep["A10"] = "온라인몰 주문 비중"
rep["B10"] = "=대시보드!B20"
rep["B7"].number_format = '"₩"#,##0'
rep["B9"].number_format = '"₩"#,##0'
rep["B10"].number_format = "0.0%"
for r in range(6, 11):
    box(rep.cell(r, 1), Font(name="Arial", bold=True, size=11), paper_fill)
    box(rep.cell(r, 2), form_f, card_fill, center)

rep["A13"] = "요약"
rep["A13"].font = Font(name="Arial", bold=True, size=13)
rep["A14"] = (
    '="완료 주문은 "&TEXT(B6,"#,##0")&"건, 순매출은 "&TEXT(B7,"₩#,##0")&"이다. '
    '온라인몰 주문은 "&TEXT(B8,"#,##0")&"건으로 전체의 "&TEXT(B10,"0.0%")&"를 차지한다."'
)
rep["A14"].font = form_f
rep["A14"].alignment = Alignment(wrap_text=True, vertical="top")
rep.merge_cells("A14:F17")
rep.row_dimensions[14].height = 48
rep["A19"] = "등급 기준은 실습 Section 4와 같다. EMP013(황성민)부터 확인해 강의 예시와 맞춰 보면 된다."
rep["A19"].font = muted_f
rep.merge_cells("A19:F19")
rep.column_dimensions["A"].width = 22
rep.column_dimensions["B"].width = 18

out = "/Users/leejaehun/Desktop/오늘마감/실습4_엑셀자동화.xlsx"
wb.save(out)
print(out, "orders", n_orders)
